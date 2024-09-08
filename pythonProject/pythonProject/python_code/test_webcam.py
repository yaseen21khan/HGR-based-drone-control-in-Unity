import csv
import time
import cv2
import joblib
import matplotlib.pyplot as plt
import mediapipe as mp
import numpy as np
import openpyxl
import pandas as pd
import seaborn as sn
import tensorflow as tf
from openpyxl.reader.excel import load_workbook
from sklearn.metrics import accuracy_score, f1_score
from openpyxl import load_workbook, Workbook
import socket

recorded_data =[]
detected_label_name = 'Unknown'
confirmation_label = ""
confrimation = False
label_mapping = {}
global_labels=[]
confidence = None
counter = 10
excel_file="output.xlsx"
label_mappings = {
    "TAKEOFF": "Take off",
    "LAND": "Land",
    "move up": "Move up",
    "move down": "Move down",
    "Advance": "Move forward",
    "backward": "Move backward",
    "move left": "Move left",
    "move right": "Move right",
    "clockwise rotation": "Turn clockwise",
    "counterclockwise rotation": "Turn anti-clockwise"
}

def copy_excel_sheet(original_file, copy_file):
    wb = load_workbook(original_file)
    ws_original = wb.active
    wb_copy = Workbook()
    ws_copy = wb_copy.active
    for row in ws_original.iter_rows(values_only=True):
        ws_copy.append(row)
    wb_copy.save(copy_file)
def append_data_in_excel(counter, excel_file):
    copy_excel_sheet(excel_file, "new_file.xlsx")
    wb = load_workbook(excel_file)
    ws = wb.active
    max_row = ws.max_row

    data_col1 = [ws.cell(row=i, column=1).value for i in range(2, max_row + 1)]
    data_col2 = [ws.cell(row=i, column=2).value for i in range(2, max_row + 1)]

    for i in range(1, counter):
        for j in range(len(data_col1)):
            # Calculate the correct row index for insertion
            row_index = max_row + (max_row - 1) * (i - 1) + j + 1
            ws.cell(row=row_index, column=1).value = data_col1[j]
            ws.cell(row=row_index, column=2).value = data_col2[j]

    wb.save(excel_file)
def count_entries_in_second_column(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
    count = 0
    for cell in ws['B']:
        if cell.value:
            count += 1
    return count
#============ files =========================================
append_data_in_excel(counter, excel_file)
number_of_rows = count_entries_in_second_column(excel_file)
number_of_rows-=1
#===========================================================

csv_path = 'label_mapping_30_unsorted.csv'
scaler = joblib.load('landmark_dataset_17_classes_model_3.pkl')
loaded_model = tf.keras.models.load_model('landmark_dataset_17_classes_model_3.hdf5')
mp_holistic = mp.solutions.holistic  # Holistic model

def plot_confusion_matrix(filename):
    model_pred = pd.read_excel(filename, index_col=[0])
    cm = pd.crosstab(model_pred.Predicted_class, model_pred.Actual_class)
    accuracy = accuracy_score(model_pred.Actual_class, model_pred.Predicted_class)
    f1 = f1_score(model_pred.Actual_class, model_pred.Predicted_class, average='weighted')
    plt.figure(figsize=(10, 8))
    sn.heatmap(cm, annot=True, cmap="Blues", xticklabels=True, yticklabels=True, cbar=True)
    plt.xticks(rotation=45)
    plt.yticks(rotation=45)
    plt.text(0.5, 1.05, f'Accuracy: {accuracy:.2f}', horizontalalignment='center', verticalalignment='center',
             transform=plt.gca().transAxes)
    plt.text(0.5, 1.1, f'F1 Score: {f1:.2f}', horizontalalignment='center', verticalalignment='center',
             transform=plt.gca().transAxes)
    plt.show()
    plt.tight_layout()
def write_to_excel(array_list, filename):
    wb = load_workbook(filename)
    ws = wb.active
    last_row = 1
    for i, value in enumerate(array_list, start=1):
        ws.cell(row=last_row + i, column=3, value=value)
    wb.save(filename)

def erase_second_column(filename):
    wb = load_workbook(filename)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.value = None
    wb.save(filename)
def get_most_frequent_label( labels):
    freq = {}
    for label in labels:
        freq[label] = freq.get(label, 0) + 1

    most_frequent_label = max(freq, key=freq.get)

    if most_frequent_label == "No hands detected in the input image.":
        least_frequent_label = min(freq, key=freq.get)
        return least_frequent_label
    else:
        return most_frequent_label
def process_frames(rgb_frame):
    rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    results = holistic.process(rgb_frame)
    landmark_list = extract_keypoints(results)

    return landmark_list

def send_label_to_server( label):
    host, port = "210.107.229.105", 25003
    # host, port = "192.168.0.8", 25003
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        sock.connect((host, port))
        sock.sendall(label.encode("utf-8"))
        response = sock.recv(1024).decode("utf-8")
        print(response)
    finally:
        sock.close()
def extract_keypoints(results):
    landmark_list = []
    if results.left_hand_landmarks and results.right_hand_landmarks:
        lh = np.array([[res.x, res.y, res.z] for res in results.left_hand_landmarks.landmark]).flatten()
        rh = np.array([[res.x, res.y, res.z] for res in results.right_hand_landmarks.landmark]).flatten()
        if lh.size > 0 and rh.size > 0:
            landmark_list = np.concatenate([lh, rh])
    return landmark_list

with open(csv_path, mode='r') as file:
    reader = csv.reader(file, delimiter='\t')  # Specify tab delimiter
    for index, row in enumerate(reader):
        label_mapping[index] = row[0]

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Hand Gestures'
sheet['A1'] = 'Ground Truth Label'
sheet['B1'] = 'Detected Label'
border_color = (0, 255, 0)
border_thickness = 5
color = "green"
cap = cv2.VideoCapture(0)
previous_mapped_label = None

with mp_holistic.Holistic(min_detection_confidence=0.5, min_tracking_confidence=0.5) as holistic:
    print("number of rows:", number_of_rows)
    row_counter = 5
    start_time = time.time()
    while cap.isOpened():
        ret, frame = cap.read()
        if not ret:
            break

        elapsed_time = time.time() - start_time
        if elapsed_time >= 2:
            start_time = time.time()
            if color == "green":
                color = "red"
            else:
                color = "green"


        if color == "green":
            border_color = (0, 0, 255)  # Red color
            if len(global_labels) > 0:
                data = get_most_frequent_label(global_labels)
                print("Added entry :", data)
                recorded_data.append(data)
                detected_label_name = data
                mapped_label = label_mappings.get(detected_label_name, detected_label_name)
                cv2.putText(frame, mapped_label, (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 3)

                if mapped_label == "Check" and previous_mapped_label is not None:
                    send_label_to_server(previous_mapped_label)
                    print("Sending previous gesture:", previous_mapped_label)
                elif mapped_label == "Check" or mapped_label == "Cancel":
                    confrimation = True
                else:
                    confrimation = False

                if confrimation:
                    if detected_label_name != "Unknown":
                        send_label_to_server(mapped_label)
                    print("Confirmation gesture")
                else:
                    print("Basic gesture")

                global_labels = []
                previous_mapped_label = mapped_label
        else:
            border_color = (0, 255, 0)  # Green color
        if color == "red":
            cv2.putText(frame, "Recording frames .. .", (100, 300), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 3)
            detected_label_name = 'Unknown'
            rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            landmark_list = process_frames(rgb_frame)
            if len(landmark_list) >= 126:
                standardized_landmarks = scaler.transform([landmark_list])
                predict_result = loaded_model.predict(standardized_landmarks)
                predicted_class_index = np.argmax(predict_result)
                confidence = np.max(predict_result)
                detected_label_name = label_mapping.get(predicted_class_index, 'Unknown')
                detected_label_name = detected_label_name.split(',')[0]
                global_labels.append(detected_label_name)
        else:
            str = f"{detected_label_name} : [{confidence}]".format(confidence)
            cv2.putText(frame,str, (100, 300), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 3)
        frame = cv2.rectangle(frame, (0, 0), (frame.shape[1], frame.shape[0]), border_color,border_thickness)
        cv2.imshow('Hand Gestures', frame)
        if len(recorded_data) >= number_of_rows:
            # write_to_excel(recorded_data, "output.xlsx")
            # plot_confusion_matrix("output.xlsx")
            # erase_second_column("output.xlsx")
            for entry in recorded_data:
                print("Entries inside recorded data are:", entry)
            break

        if cv2.waitKey(1) & 0xFF == 27:
            # write_to_excel(recorded_data, "output.xlsx")
            # plot_confusion_matrix("output.xlsx")
            # erase_second_column("output.xlsx")
            for entry in recorded_data:
                print("Entries inside recorded data are:", entry)
            break

cap.release()
cv2.destroyAllWindows()
